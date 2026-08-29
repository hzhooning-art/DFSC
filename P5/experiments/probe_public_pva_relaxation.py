"""Stage 62: direct public-data audit of PVA stress-relaxation curves."""

from __future__ import annotations

import hashlib
import json
import math
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import least_squares, nnls
from scipy.stats import bootstrap, wilcoxon


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "external" / "pva_gpe_zenodo_21333840" / "Stress_relaxation_data.xlsx"
RESULTS = ROOT / "results"
OUTPUT_JSON = RESULTS / "public_pva_relaxation.json"
OUTPUT_MD = RESULTS / "public_pva_relaxation.md"
EXPECTED_MD5 = "403b3254288b4ce8aa36cada05c0e1e4"
DOI = "10.5281/zenodo.21333840"
RANKS = (1, 2, 3)
HORIZONS = (4.0, 8.0, 15.0, 28.0)
SAMPLE_BUDGETS = (12, 24, 48, 96)
RATE_BOUNDS = (1.0 / 120.0, 20.0)
CALIBRATION_FRACTION = 0.60
DELTA_BIC_GATE = 10.0
PREDICTIVE_GAIN_GATE = 0.05
LOG_RATE_STD_GATE = 0.50
RATE_RATIO_GATE = 1.25


@dataclass(frozen=True)
class Curve:
    sample: int
    cycle: int
    time: np.ndarray
    value: np.ndarray


def file_md5(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@lru_cache(maxsize=1)
def _load_curves_cached(path: Path) -> tuple[Curve, ...]:
    if file_md5(path) != EXPECTED_MD5:
        raise ValueError("public workbook checksum does not match the frozen source")
    workbook = load_workbook(path, read_only=True, data_only=True)
    curves = []
    for sample in (1, 2, 3):
        sheet = workbook[f"Sample {sample}"]
        rows = list(sheet.iter_rows(min_row=4, max_row=min(sheet.max_row, 2000), values_only=True))
        for cycle, (time_col, displacement_col, force_col) in enumerate(
            ((3, 4, 5), (8, 9, 10), (13, 14, 15)), start=1
        ):
            def column(index: int) -> np.ndarray:
                return np.asarray([
                    float(row[index - 1]) if len(row) >= index and isinstance(row[index - 1], (int, float)) else np.nan
                    for row in rows
                ], dtype=float)

            time = column(time_col)
            displacement = column(displacement_col)
            force = column(force_col)
            valid = np.isfinite(time) & np.isfinite(displacement) & np.isfinite(force)
            time, displacement, force = time[valid], displacement[valid], force[valid]
            threshold = 0.999 * float(np.max(displacement))
            transition = int(np.flatnonzero(displacement >= threshold)[0])
            start_candidates = np.flatnonzero(time >= time[transition] + 0.25)
            start = int(start_candidates[0]) if len(start_candidates) else transition
            hold_time = time[start:] - time[start]
            hold_force = force[start:] / force[start]
            curves.append(Curve(sample, cycle, hold_time, hold_force))
    workbook.close()
    if len(curves) != 9:
        raise ValueError("expected three cycles for each of three specimens")
    return tuple(curves)


def load_curves(path: Path = DATA) -> list[Curve]:
    return list(_load_curves_cached(path))


def resample(curve: Curve, horizon: float, budget: int) -> Curve:
    end = min(float(curve.time[-1]), float(horizon))
    if end <= 0.5:
        raise ValueError("declared horizon is too short")
    grid = np.linspace(0.0, end, int(budget))
    return Curve(curve.sample, curve.cycle, grid, np.interp(grid, curve.time, curve.value))


def _design(time: np.ndarray, rates: np.ndarray) -> np.ndarray:
    return np.column_stack((np.ones_like(time), np.exp(-np.outer(time, rates))))


def fit_coefficients(time: np.ndarray, value: np.ndarray, rates: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    coefficients, _ = nnls(_design(time, rates), value)
    prediction = _design(time, rates) @ coefficients
    return coefficients, prediction


def _initial_rates(rank: int, start: int) -> np.ndarray:
    low, high = RATE_BOUNDS
    shifts = (0.55, 0.8, 1.0, 1.35, 1.8, 2.4)
    base = np.geomspace(0.025, 2.5, rank) * shifts[start % len(shifts)]
    return np.clip(base, low * 1.01, high / 1.01)


def fit_shared_rates(curves: list[Curve], rank: int, starts: int = 6) -> dict:
    low, high = np.log(RATE_BOUNDS)

    def residual(log_rates: np.ndarray) -> np.ndarray:
        rates = np.sort(np.exp(log_rates))
        blocks = []
        for curve in curves:
            _, prediction = fit_coefficients(curve.time, curve.value, rates)
            blocks.append(prediction - curve.value)
        return np.concatenate(blocks)

    best = None
    for start in range(starts):
        result = least_squares(
            residual,
            np.log(_initial_rates(rank, start)),
            bounds=(np.full(rank, low), np.full(rank, high)),
            max_nfev=500,
            ftol=1e-11,
            xtol=1e-11,
            gtol=1e-11,
        )
        rates = np.sort(np.exp(result.x))
        errors = residual(np.log(rates))
        sse = float(errors @ errors)
        if best is None or sse < best["sse"]:
            best = {"rates": rates, "sse": sse, "success": bool(result.success)}
    assert best is not None
    n = sum(len(curve.time) for curve in curves)
    parameters = rank + len(curves) * (rank + 1)
    best["bic"] = float(n * np.log(max(best["sse"] / n, 1e-300)) + parameters * np.log(n))
    ratios = best["rates"][1:] / best["rates"][:-1]
    best["minimum_rate_ratio"] = float(np.min(ratios)) if len(ratios) else math.inf
    return best


def held_prediction_errors(curves: list[Curve], held_sample: int, rates: np.ndarray) -> list[float]:
    errors = []
    for curve in curves:
        if curve.sample != held_sample:
            continue
        split = max(4, int(math.ceil(CALIBRATION_FRACTION * len(curve.time))))
        coefficients, _ = fit_coefficients(curve.time[:split], curve.value[:split], rates)
        prediction = _design(curve.time[split:], rates) @ coefficients
        scale = max(float(np.ptp(curve.value[:split])), 0.05)
        errors.append(float(np.sqrt(np.mean((prediction - curve.value[split:]) ** 2)) / scale))
    return errors


def evaluate_grid_cell(curves: list[Curve], horizon: float, budget: int, starts: int = 6) -> dict:
    sampled = [resample(curve, horizon, budget) for curve in curves]
    rank_records = {}
    for rank in RANKS:
        folds = []
        for held_sample in (1, 2, 3):
            training = [curve for curve in sampled if curve.sample != held_sample]
            fit = fit_shared_rates(training, rank, starts=starts)
            folds.append({
                "held_sample": held_sample,
                "rates": fit["rates"].tolist(),
                "bic": fit["bic"],
                "sse": fit["sse"],
                "minimum_rate_ratio": fit["minimum_rate_ratio"],
                "errors": held_prediction_errors(sampled, held_sample, fit["rates"]),
                "success": fit["success"],
            })
        log_rates = np.asarray([np.log(row["rates"]) for row in folds])
        curve_errors = [value for row in folds for value in row["errors"]]
        rank_records[str(rank)] = {
            "folds": folds,
            "mean_bic": float(np.mean([row["bic"] for row in folds])),
            "median_prediction_nrmse": float(np.median(curve_errors)),
            "curve_prediction_nrmse": curve_errors,
            "max_log_rate_std": float(np.max(np.std(log_rates, axis=0, ddof=1))),
            "minimum_rate_ratio": float(min(row["minimum_rate_ratio"] for row in folds)),
            "all_finite": bool(all(row["success"] for row in folds) and np.isfinite(log_rates).all()),
        }

    selected = 1
    conflict = False
    transitions = []
    for rank in (2, 3):
        lower = rank_records[str(rank - 1)]
        current = rank_records[str(rank)]
        delta_bic = lower["mean_bic"] - current["mean_bic"]
        gain = (lower["median_prediction_nrmse"] - current["median_prediction_nrmse"]) / max(
            lower["median_prediction_nrmse"], 1e-15
        )
        gates = {
            "bic": bool(delta_bic >= DELTA_BIC_GATE),
            "prediction": bool(gain >= PREDICTIVE_GAIN_GATE),
            "stability": bool(current["max_log_rate_std"] <= LOG_RATE_STD_GATE),
            "separation": bool(current["minimum_rate_ratio"] >= RATE_RATIO_GATE),
            "finite": current["all_finite"],
        }
        transitions.append({"to_rank": rank, "delta_bic": delta_bic, "prediction_gain": gain, "gates": gates})
        if gates["bic"]:
            if all(gates.values()) and selected == rank - 1:
                selected = rank
            else:
                conflict = True
    decision = "INDETERMINATE" if conflict else f"SUPPORTED_RANK_{selected}"
    return {
        "horizon_seconds": horizon,
        "samples_per_curve": budget,
        "rank_records": rank_records,
        "transitions": transitions,
        "decision": decision,
    }


def _paired_statistics(cell: dict) -> dict:
    baseline = np.asarray(cell["rank_records"]["1"]["curve_prediction_nrmse"])
    records = {}
    for rank in (2, 3):
        candidate = np.asarray(cell["rank_records"][str(rank)]["curve_prediction_nrmse"])
        improvement = (baseline - candidate) / np.maximum(baseline, 1e-15)
        interval = bootstrap(
            (improvement,), np.median, confidence_level=0.95,
            n_resamples=5000, random_state=np.random.default_rng(6200 + rank), method="percentile",
        ).confidence_interval
        test = wilcoxon(baseline, candidate, alternative="greater", method="auto")
        records[str(rank)] = {
            "median_relative_improvement": float(np.median(improvement)),
            "bootstrap_95pct": [float(interval.low), float(interval.high)],
            "wilcoxon_statistic": float(test.statistic),
            "wilcoxon_p_one_sided": float(test.pvalue),
            "n_curves": int(len(improvement)),
        }
    return records


def build_payload() -> dict:
    curves = load_curves()
    full = evaluate_grid_cell(curves, 28.0, 96, starts=8)
    boundary = []
    for horizon in HORIZONS:
        for budget in SAMPLE_BUDGETS:
            boundary.append(evaluate_grid_cell(curves, horizon, budget, starts=4))
            print(f"stage62 horizon={horizon:g} budget={budget} decision={boundary[-1]['decision']}", flush=True)
    source = {
        "title": "Stress Relaxation Test Dataset of Cylindrical PVA Gel Polymer Electrolyte (GPE) Samples",
        "doi": DOI,
        "record_url": "https://zenodo.org/records/21333840",
        "access": "open",
        "license": "CC-BY-4.0",
        "license_metadata_field": "metadata.license.id",
        "file": str(DATA),
        "md5": file_md5(DATA),
        "specimens": 3,
        "cycles_per_specimen": 3,
    }
    checks = {
        "source_checksum": source["md5"] == EXPECTED_MD5,
        "nine_direct_observed_curves": len(curves) == 9,
        "all_boundary_cells_complete": len(boundary) == len(HORIZONS) * len(SAMPLE_BUDGETS),
        "all_decisions_declared": all(row["decision"].startswith(("SUPPORTED_RANK_", "INDETERMINATE")) for row in boundary),
    }
    return {
        "experiment": "stage62_direct_public_pva_relaxation",
        "protocol_frozen_before_fit": True,
        "source": source,
        "preprocessing": {"displacement_fraction": 0.999, "transition_skip_seconds": 0.25, "smoothing": "none"},
        "model": "nonnegative shared-rate exponential realization with curve-specific amplitudes and baseline",
        "gates": {
            "delta_bic": DELTA_BIC_GATE,
            "relative_prediction_gain": PREDICTIVE_GAIN_GATE,
            "maximum_log_rate_std": LOG_RATE_STD_GATE,
            "minimum_adjacent_rate_ratio": RATE_RATIO_GATE,
        },
        "full_task": full,
        "paired_statistics": _paired_statistics(full),
        "boundary": boundary,
        "checks": checks,
        "route_pass": all(checks.values()),
        "claim_boundary": "A supported rank is an empirical shared relaxation realization on the declared windows, not proof of a unique molecular mechanism.",
    }


def write_outputs(payload: dict) -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    full = payload["full_task"]
    lines = [
        "# Direct public PVA relaxation audit", "",
        f"Route pass: **{payload['route_pass']}**.", "",
        f"Decision at 28 s / 96 samples per curve: **{full['decision']}**.", "",
        "## Rank diagnostics", "",
        "| Rank | Mean BIC | Median held-sample NRMSE | Max log-rate std | Min rate ratio |",
        "|---:|---:|---:|---:|---:|",
    ]
    for rank in RANKS:
        row = full["rank_records"][str(rank)]
        lines.append(
            f"| {rank} | {row['mean_bic']:.3f} | {row['median_prediction_nrmse']:.5f} | "
            f"{row['max_log_rate_std']:.4f} | {row['minimum_rate_ratio']:.3f} |"
        )
    lines.extend(["", "The observations are fitted directly; no synthetic residual injection is used."])
    OUTPUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    payload = build_payload()
    write_outputs(payload)
    print(json.dumps({"decision": payload["full_task"]["decision"], "route_pass": payload["route_pass"]}, indent=2))


if __name__ == "__main__":
    main()
