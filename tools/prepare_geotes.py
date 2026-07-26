"""Normalize the public GeoTES thermocouple workbook for dfsc experiments."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path


DOI = "10.5281/zenodo.18979098"
SOURCE = f"https://doi.org/{DOI}"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, default=None)
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required only for this one-time preprocessing step") from exc

    workbook = args.workbook.resolve()
    output_dir = (args.output_dir or workbook.parent).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    book = openpyxl.load_workbook(workbook, data_only=True, read_only=True)

    channels: dict[str, dict[object, float]] = {}
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        label = str(sheet.cell(1, 2).value)
        channels[label] = {
            timestamp: float(value)
            for timestamp, value in rows
            if timestamp is not None and value is not None
        }

    timestamps = sorted(set.intersection(*(set(values) for values in channels.values())))
    if len(timestamps) < 100:
        raise ValueError("too few common timestamps across thermocouple channels")

    csv_path = output_dir / "geotes_thermocouples.csv"
    labels = sorted(channels)
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["timestamp", "elapsed_hours", *[f"{label}_celsius" for label in labels]])
        origin = timestamps[0]
        for timestamp in timestamps:
            elapsed = (timestamp - origin).total_seconds() / 3600.0
            writer.writerow([timestamp.isoformat(), f"{elapsed:.9f}", *[channels[label][timestamp] for label in labels]])

    manifest = {
        "name": "GeoTES pilot-scale thermocouple histories",
        "domain": "high-temperature thermal energy storage",
        "task": "cross-cycle forced multi-channel response identification",
        "source": SOURCE,
        "license": "CC BY 4.0",
        "citation": "Georgiou, Dataset for: Modular High-Temperature Geopolymer Thermal Energy Storage, Zenodo (2026)",
        "splits": {
            "train": "first charging-discharge cycle",
            "test": "second charging-discharge cycle",
        },
        "tensors": {},
        "files": {
            workbook.name: {"sha256": sha256(workbook), "role": "unaltered source workbook"},
            csv_path.name: {"sha256": sha256(csv_path), "role": "timestamp-aligned derived table"},
        },
        "measurement_boundary": "T1 is used as the measured driving channel; T2-T4 are response channels. No undocumented sensor coordinates are assumed.",
        "preprocessing": "inner timestamp join across T1-T4; Celsius values retained; no interpolation or smoothing",
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps({"rows": len(timestamps), "csv": str(csv_path), "manifest": str(output_dir / 'manifest.json')}, indent=2))


if __name__ == "__main__":
    main()
