"""Convert the Zenodo heated-steam workbook into a long-form CSV."""

from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "external" / "heated_steam"
SOURCE = DATA / "2025_Garcia-Martinez_Summary_LabNumData.xlsx"
EXPECTED_MD5 = "d14aa1ee2a77890e75817ef0e185a363"


def main() -> None:
    digest = hashlib.md5(SOURCE.read_bytes()).hexdigest()
    if digest != EXPECTED_MD5:
        raise RuntimeError(f"source MD5 mismatch: {digest}")
    workbook = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    summary = workbook["Summary"]
    summary_rows = list(summary.iter_rows(min_row=14, max_row=29, max_col=5, values_only=True))
    conditions = {
        experiment: tuple(summary_rows[experiment - 1][1:5])
        for experiment in range(1, 17)
    }
    sheet = workbook["Temperature_Profiles"]
    sheet_rows = sheet.iter_rows(values_only=True)
    experiment_headers = next(sheet_rows)
    column_headers = next(sheet_rows)
    rows = []
    depths_by_experiment = {}
    for experiment in range(1, 17):
        start = 1 + 11 * (experiment - 1)
        headers = column_headers[start - 1 : start + 10]
        depths = [float(str(header).split("Depth_")[1].split("m_")[0]) for header in headers[1:]]
        depths_by_experiment[experiment] = depths
    for values in sheet_rows:
        for experiment in range(1, 17):
            start = 11 * (experiment - 1)
            time_h = values[start]
            if time_h is None:
                continue
            depths = depths_by_experiment[experiment]
            flow, inlet_temperature, column_height, vwc = conditions[experiment]
            for offset, depth in enumerate(depths, 1):
                temperature = values[start + offset]
                if temperature is None:
                    continue
                rows.append(
                    {
                        "experiment": experiment,
                        "time_h": float(time_h),
                        "depth_m": depth,
                        "temperature_k": float(temperature),
                        "inflow_1e-5_kg_s": float(flow),
                        "inlet_temperature_k": float(inlet_temperature),
                        "column_height_m": float(column_height),
                        "vwc": float(vwc),
                    }
                )
    target = DATA / "heated_steam_profiles.csv"
    with target.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    manifest = {
        "name": "Heated steam injection experimental temperature profiles",
        "source_doi": "10.5281/zenodo.15064388",
        "source_file": SOURCE.name,
        "source_md5": digest,
        "license": "CC BY 4.0",
        "creators": ["Noe Garcia-Martinez", "Tarsilo Girona", "David Benavente"],
        "experiments": 16,
        "depth_coordinates_m": depths_by_experiment[1],
        "rows": len(rows),
        "conversion": "Temperature_Profiles worksheet converted without interpolation or smoothing.",
    }
    (DATA / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
